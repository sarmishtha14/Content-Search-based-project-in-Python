from tkinter import *
import pandas as pd
import numpy as np
import openpyxl as opx 
import re
import csv
import os
import random
from nltk.stem import WordNetLemmatizer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import nltk
import operator

Labelfont= ('Helvetica',11,'bold')
stop_words = set(stopwords.words('english')) 
fields = 'Enter Invoice Number', 'Enter Text'
lemmatizer = WordNetLemmatizer()
rootdir ='C:\\Users\\Mahe\\Desktop\\Python\\ExcelFiles'
dutyfee = ["5%","Free","10%","20%","15%"]
info = ""
code = ""
maxi = 0

#Fetches data from ui
def fetch(entries):    
        root1 = Tk()
        root1.title("Output")
        field1 = fields[0]
        inv = entries[0].get()
        field2 = fields[1]
        text  = entries[1].get()
        text = text.lower()
        word_tokens = word_tokenize(text)
        print('%s: "%s"' % (field1, inv))
        print('%s: "%s"' % (field2, text))
        lst = []
        for w in word_tokens:
            if w not in stop_words:
                lst.append(w)
        for item in range(len(lst)):
            lst[item] = lemmatizer.lemmatize(lst[item])
        counts = dict()
        len1 = len(lst)
        for name in lst:  
            counts[name]=counts.get(name,0)+1
        count2 = 0
        for item in counts:
            count2 = count2 + 1
        found2 = 0
        cnt = 0
            
        #Function to create tables in excel
        def createTable(dutyfee):
            fname = "chapter_10_final1.xlsx" 
            f=open(fname,errors="ignore")
            wb = opx.load_workbook(fname)  
            sname = "Sheet"
            sheet = workBook.get_sheet_by_name(sname)
            maxRows = sheet.max_row

            f = 'DB_transactions.xlsx'
            workBook = opx.load_workbook(f)  
            sname1 = "Sheet2"        

            #To create a sheet if it doesnt exist
            if sname1 not in workBook.sheetnames:
                sheet1 = workBook.create_sheet(sname1)
            else:
                sheet1 = workBook.get_sheet_by_name(sname1)    
            row2 = sheet1.max_row
            count = 1
   
            if sheet1.cell(row = 1,column = 20).value == None:
                row1 = 0
            else:
                row1 = sheet1.cell(row = 1,column = 20).value
                
            sheet1.cell(row = 1,column = 1).value = 'CHAPTER'
            sheet1.cell(row = 1,column = 2).value = 'TEXT'
            sheet1.cell(row = 1,column = 3).value = 'ROW_REF'
            sheet1.cell(row = 1,column = 4).value = 'HTC_CODE'
            sheet1.cell(row = 1,column = 7).value = 'CHAPTER'
            sheet1.cell(row = 1,column = 8).value = 'CATEGORY'
            sheet1.cell(row = 1,column = 9).value = 'ROW_REF'
            sheet1.cell(row = 1,column = 11).value = 'HTC_CODE'
            sheet1.cell(row = 1,column = 12).value = 'DUTY_FEE'
            sheet1.cell(row = 1,column = 13).value = 'DUTY_FEE'
            col = row1
            k = row1 + 1
            for i in range(2,maxRows+1):
                k = k+1
                read = sheet.cell(row=i,column = 2).value
                cat = read[0]
                for j in range(1,len(read)-1):
                    if read[j]>='A' and read[j]<='Z':
                        break
                    else:
                        cat = cat + read[j]
                    
                ref = sheet.cell(row = i,column = 1).value
                code = sheet.cell(row = i,column = 3).value
                chap = re.findall(".+[0-9]", fname)
                sheet1.cell(row = i+row2-1,column = 1).value = chap[0]
                sheet1.cell(row = i+row2-1,column = 2).value = read
                sheet1.cell(row = i+row2-1,column = 3).value = ref+row2
                sheet1.cell(row = i+row2-1,column = 4).value = code
                sheet1.cell(row = i+row2-1,column = 11).value = code
                sheet1.cell(row = i+row2-1,column = 12).value = random.choice(dutyfee)
                sheet1.cell(row = i+row2-1,column = 13).value = random.choice(dutyfee)

                if sheet1.cell(row = k-1,column =8).value == cat:
                    k = k-1
                    count = count + 1
                else:
                    sheet1.cell(row = k, column = 8).value = cat
                    sheet1.cell(row = k,column = 9).value = count + row2
                    sheet1.cell(row = k,column = 7).value = chap[0]
                    
                    col = col + 1
                    count = count + 1
            sheet1.cell(row = 1,column = 20).value = col
            workBook.save(f)

        #Stores input and output data in an excel file        
        def storeData(inv,text,info,code,flag,dutyfee1,dutyfee2):
            wb = opx.load_workbook('DB_transactions.xlsx')
            ws = wb.get_sheet_by_name('Sheet1')
            df = pd.read_excel('DB_transactions.xlsx')
            row1 = max([cell[0] for cell in ws._cells if cell[1] == 1])   
            ws.cell(row=row1+1, column=1).value = inv
            ws.cell(row=row1+1, column=2).value = text
            ws.cell(row=row1+1, column=4).value = inv
            ws.cell(row=row1+1, column=5).value = text
            ws.cell(row=row1+1, column=6).value = code
            ws.cell(row=row1+1, column=7).value = info
            ws.cell(row=row1+1, column=8).value = dutyfee1
            ws.cell(row=row1+1, column=9).value = dutyfee2
            ws.cell(row=row1+1, column=10).value = flag
            wb.save('DB_transactions.xlsx')
                    
        #Printing output on ui
        def printio(root1,info,code):
                
            #container = tk.Canvas(master_frame, height=300, width=720, scrollregion=(0,0,300,720))
            #vbar = Scrollbar(orient="vertical")
            lab1 = Label(root1, width= 100, text = "Information: "+info)
            lab2 = Label(root1, width= 30, text = "Code: "+code)
            lab1.config(bg='pink', fg='magenta')  
            lab1.config(font=Labelfont)           
            lab1.config(height=1, width=50)  
            lab2.config(bg='pink', fg='magenta')  
            lab2.config(font=Labelfont)           
            lab2.config(height=1, width=50)  
            lab1.pack(side=TOP,expand = True,fill = BOTH)
            lab2.pack(side=TOP,expand = True,fill = BOTH)
                    
        #Searching for some text in the content
        def searchAlgorithm(counts,count2,len1,lst,inv,text,cnt):
            fname = "DB_transactions.xlsx" 
            f=open(fname,errors="ignore")
            workBook = opx.load_workbook(fname)  
            sname = "Sheet2"
            sheet1 = workBook.get_sheet_by_name(sname)
            maxRows = sheet1.max_row            
            
            col = sheet1.cell(row = 1,column = 20).value
            found = 0
            r1 = list()
            r2 = list()
            for i in range(2,col+2):
                count = 0
                read = sheet1.cell(row = i, column = 8).value
                read = read.lower()
                l = read.split()
                for item in range(len(l)):
                    l[item] =lemmatizer.lemmatize(l[item])
                    for j in lst:
                        for k in l:
                            if j == k:
                                count = count + 1
                    if count>=1:
                        found = 1
                        r1.append(sheet1.cell(row = i,column = 9).value)
                        if i == col+1:
                            r2.append(maxRows)
                        else:
                            r2.append(sheet1.cell(row = i+1,column = 9).value)
                if found == 0:
                    r1.append(1)
                    r2.append(maxRows)
                for a in range(len(r1)):
                    ref = r1[a]
                    ref2 = r2[a]
                    found2 = 0
                    for i in range(ref,ref2):
                        read1 = sheet1.cell(row = i+1, column = 2).value
                        counts2 = dict() 
                        read2 = read1.lower()
                        lst2 = read2.split()
                        for item in range(len(lst2)):
                            lst2[item] =lemmatizer.lemmatize(lst2[item])
                        for j in range(len1):
                            for k in range(len(lst2)):
                                if lst[j]==lst2[k] and (lst[j]=='not' or lst[j]=='no'):
                                    counts2[lst[j]]=counts2.get(lst[j],0)+1
                                elif lst[j] == lst2[k] and (lst[j-1]!='not' and lst[j-1]!='no'):
                                    if lst2[k-1] != 'not' and lst2[k-1]!='no':                        
                                        counts2[lst[j]]=counts2.get(lst[j],0)+1
                                    else:
                                        continue
                                elif lst[j]==lst2[k] and (lst[j-1]=='not' or lst[j-1]=='no'):
                                    counts2[lst[j]]=counts2.get(lst[j],0)+1
                                    
                        f = 0
                        count3 = 0
                        found2 = 0
                        for item in counts2:
                            count3 = count3 + 1
                        if count3 == count2:
                            for word,count3 in counts2.items():
                                for word1,count4 in counts.items():
                                    if word==word1:
                                        f = 1
                                        if count3 > 0:
                                            found2 = 1
                                        else:
                                            found2 = 0
                                            break
                                if f == 1:
                                    continue
                                else:
                                    break
                            if found2 == 0:
                                continue
                            elif found2 == 1:
                                code = sheet1.cell(row = i+1,column = 4).value
                                info = sheet1.cell(row = i+1,column = 2).value
                                dutyfee1 = sheet1.cell(row = i+1,column = 12).value
                                dutyfee2 = sheet1.cell(row = i+1,column = 13).value
                                print(info)
                                print(code)
                                cnt = cnt + 1
                                printio(root1,info,code)
                                #storeData(inv,text,info,code,found2,dutyfee1,dutyfee2)
                return cnt
        
        #Cleaning of data in the excel file
        def cleanData(maxRows,wb,sheet,fname):
            for i in range(2,maxRows):
                for j in range(2,maxRows-i):
                    if sheet.cell(row = j,column = 2).value > sheet.cell(row =j+1, column = 2).value:
                        temp = sheet.cell(row = j,column = 2).value
                        sheet.cell(row = j,column = 2).value = sheet.cell(row = j+1,column = 2).value
                        sheet.cell(row = j+1,column = 2).value= temp
            for i in range(2,maxRows+1):
                read = sheet.cell(row = i, column = 2).value       
                line = re.findall("[a-zA-Z]+", read)
                li = list()
                for j in line:
                    if j != 'NA' and j!='i':
                        li.append(j)
                lin = " ".join(li)
                num = ".".join(re.findall("[0-9]+",read))
                sheet.cell(row = i, column = 2).value = lin
                sheet.cell(row = i,column = 3).value = num
        
            wb.save(fname)
                
        #Function calling
        #cleanData(maxRows,workBook,sheet,fname)
        #createTable(dutyfee)
        #combineData(workBook,sheet1,maxRows,fname,rootdir)
        found2 = searchAlgorithm(counts,count2,len1,lst,inv,text,cnt)
        cnt = found2
        if found2 == None or found2 == 0:
            found2 = 0
        else:
            found2 = 1
        
        if found2 == 0:
            info = "Text not found"
            code = "NA"
            dutyfee1 = "NA"
            dutyfee2 = "NA"
            print(info)
            printio(root1,info,code)
            #storeData(inv,text,info,code,found2,dutyfee1,dutyfee2)
        root1.mainloop()
         
            
#Makes the form
def makeform(root, fields,info,code):
    
        entries = []
        for field in fields: 
            row = Frame(root, bg = "pink")
            lab = Label(row, width=20, text=field, anchor='w')
            lab.config(bg='pink', fg='magenta')  
            lab.config(font=Labelfont)           
            lab.config(height=3, width=20)
            ent = Entry(row)
            ent.config(bg='grey', fg='white')  
            ent.config(font=Labelfont)           
            ent.config(width=20)
            row.pack(side=TOP, fill=X, padx=5, pady=5)
            lab.pack(side=LEFT,expand = YES,fill = X)
            ent.pack(side=RIGHT, expand=YES, fill=X)
            entries.append(ent)
            print(entries)
        return entries

#Main program     
if __name__ == '__main__':
   root = Tk()
   root.configure(background='pink')
   ents = makeform(root, fields,info,code)
   root.bind('<Return>', (lambda event, e=ents: fetch(e)))   
   b1 = Button(root, text='Submit',
          command=(lambda e=ents: fetch(e)))
   b1.config(bg='grey', fg='white')  
   b1.config(font=Labelfont)           
   b1.config(height=1, width=10)
   b1.pack(side=LEFT, padx=60, pady=5)
   b2 = Button(root, text='Quit', command=root.quit)
   b2.config(bg='grey', fg='white')  
   b2.config(font=Labelfont)           
   b2.config(height=1, width=10)
   b2.pack(side=LEFT, padx=1, pady=5)
   root.mainloop()